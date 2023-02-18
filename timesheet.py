import openpyxl
import calendar
import json
import http.client
import sys
import glob, os
from datetime import date,datetime, timedelta

#Input arg
monthIput = sys.argv[1] if len(sys.argv) == 3 else ""
yearInput = sys.argv[2] if len(sys.argv) == 3 else ""

conn = http.client.HTTPSConnection("apigw1.bot.or.th")
headers = {
    'X-IBM-Client-Id': "d872355e-4bae-4f1b-9c59-6f0863e00d30",
    'accept': "application/json"
}

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Get the current month
today = date.today()
month = monthIput if monthIput != "" else today.strftime("%m")
year = yearInput if yearInput != "" else today.strftime("%Y")

# Call API thai holidays
conn.request("GET", "/bot/public/financial-institutions-holidays/?year="+year, headers=headers)
res = conn.getresponse().read().decode("utf-8")
data = json.loads(res)

holidays = data['result']['data']

# Convert the date strings in the holiday data to date objects
holiday_dates = [datetime.strptime(holiday['Date'], "%Y-%m-%d").date() for holiday in holidays]

# Add a new sheet with the name of the current month
sheet = workbook.active
sheet.title = "Timesheet_"+month+year

# Define the headers
headers = ["Day", "Time In", "Time Out", "OT Start", "OT Finish", "Manager Approve", "Detail", "Remark"]

# Add the headers to the first row
for i, header in enumerate(headers):
    sheet.cell(row=1, column=i+1, value=header)

# Get the number of days in the current month
num_days = calendar.monthrange(today.year, today.month)[1]

current_day = today.replace(day=1)
for i in range(num_days):
    date_to_check = datetime(year=int(year),month=int(month),day=i+1).date()
    
    weekday = date_to_check.weekday()

    sheet.cell(row=i+2, column=1, value=current_day.strftime("%d"))
    if date_to_check in holiday_dates:
        holiday = [holiday for holiday in holidays if holiday['Date'] == date_to_check.strftime("%Y-%m-%d")]
        sheet.cell(row=i+2, column=2, value=holiday[0]['HolidayDescriptionThai'])
    else:
        if weekday == 5:
            sheet.cell(row=i+2, column=2, value="--------------------------------------------------------------------------------------- วันเสาร์  -------------------------------------------------------------------------------------")
        elif weekday == 6:
            sheet.cell(row=i+2, column=2, value="-------------------------------------------------------------------------------------- วันอาทิตย์  -----------------------------------------------------------------------------------")
        else:
            sheet.cell(row=i+2, column=2, value="08:30")
            sheet.cell(row=i+2, column=3, value="17:30")
            
    current_day = current_day + timedelta(days=1)

#Delete all file .xlsx
for f in glob.glob("*.xlsx"):
    os.remove(f)

# Save the workbook to a file
workbook.save(f"Timesheet_{month}{year}.xlsx")

